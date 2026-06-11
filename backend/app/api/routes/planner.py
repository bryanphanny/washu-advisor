from io import BytesIO
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from pydantic import BaseModel
from openpyxl import load_workbook
from app.database import get_db
from app.api.deps import get_current_user

router = APIRouter()

DEFAULT_SEMESTERS = [
    {"term": "Fall",   "year": 2024, "is_completed": False},
    {"term": "Spring", "year": 2025, "is_completed": False},
    {"term": "Fall",   "year": 2025, "is_completed": False},
    {"term": "Spring", "year": 2026, "is_completed": False},
    {"term": "Fall",   "year": 2026, "is_completed": False},
    {"term": "Spring", "year": 2027, "is_completed": False},
    {"term": "Fall",   "year": 2027, "is_completed": False},
    {"term": "Spring", "year": 2028, "is_completed": False},
]


class PlannedCourse(BaseModel):
    semester_id: int
    course_code: str
    course_name: str
    credits: float


class GradeUpdate(BaseModel):
    grade: str | None = None
    grade_points: float | None = None


class APCreditCourse(BaseModel):
    course_code: str
    course_name: str
    credits: float


@router.post("/init")
def init_semesters(current_user_id: str = Depends(get_current_user)):
    """Create default semesters for a new user if they have none yet."""
    db = get_db()
    existing = db.table("semesters").select("id").eq("user_id", current_user_id).execute().data
    if existing:
        return {"status": "already_initialized", "count": len(existing)}
    rows = [{**s, "user_id": current_user_id} for s in DEFAULT_SEMESTERS]
    db.table("semesters").insert(rows).execute()
    return {"status": "initialized", "count": len(rows)}


@router.get("/semesters")
def get_semesters(current_user_id: str = Depends(get_current_user)):
    db = get_db()
    return db.table("semesters").select("*").eq("user_id", current_user_id).order("id").execute().data


@router.get("/")
def get_planned_courses(current_user_id: str = Depends(get_current_user)):
    db = get_db()
    return db.table("user_courses").select(
        "*, semesters(term, year)"
    ).eq("user_id", current_user_id).eq("is_planned", True).execute().data


@router.get("/all")
def get_all_user_courses(current_user_id: str = Depends(get_current_user)):
    db = get_db()
    return db.table("user_courses").select(
        "*, semesters(term, year)"
    ).eq("user_id", current_user_id).order("semester_id").execute().data


@router.post("/")
def add_planned_course(course: PlannedCourse, current_user_id: str = Depends(get_current_user)):
    db = get_db()
    result = db.table("user_courses").insert({
        **course.model_dump(),
        "is_planned": True,
        "user_id": current_user_id,
    }).execute()
    return result.data[0]


@router.post("/ap-credit")
def add_ap_credit(course: APCreditCourse, current_user_id: str = Depends(get_current_user)):
    db = get_db()
    result = db.table("user_courses").insert({
        **course.model_dump(),
        "is_planned": False,
        "is_transfer": True,
        "grade": "CR",
        "user_id": current_user_id,
    }).execute()
    return result.data[0]


@router.delete("/ap-credit/{course_id}")
def remove_ap_credit(course_id: int, current_user_id: str = Depends(get_current_user)):
    db = get_db()
    row = db.table("user_courses").select("*").eq("id", course_id).eq("user_id", current_user_id).eq("is_transfer", True).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="AP credit not found")
    db.table("user_courses").delete().eq("id", course_id).execute()
    return {"deleted": course_id}


@router.delete("/{course_id}")
def remove_planned_course(course_id: int, current_user_id: str = Depends(get_current_user)):
    db = get_db()
    row = db.table("user_courses").select("*").eq("id", course_id).eq("user_id", current_user_id).eq("is_planned", True).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="Planned course not found")
    db.table("user_courses").delete().eq("id", course_id).execute()
    return {"deleted": course_id}


@router.patch("/{course_id}/grade")
def set_course_grade(course_id: int, body: GradeUpdate, current_user_id: str = Depends(get_current_user)):
    db = get_db()
    result = db.table("user_courses").update({
        "grade": body.grade,
        "grade_points": body.grade_points,
    }).eq("id", course_id).eq("user_id", current_user_id).execute()
    return result.data[0]


@router.patch("/{course_id}/move")
def move_course(course_id: int, semester_id: int, current_user_id: str = Depends(get_current_user)):
    db = get_db()
    result = db.table("user_courses").update(
        {"semester_id": semester_id}
    ).eq("id", course_id).eq("user_id", current_user_id).execute()
    return result.data[0]


# ---------------------------------------------------------------------------
# Transcript upload
# ---------------------------------------------------------------------------

GRADE_POINTS_MAP = {
    "A+": 4.0, "A": 4.0, "A-": 3.7,
    "B+": 3.3, "B": 3.0, "B-": 2.7,
    "C+": 2.3, "C": 2.0, "C-": 1.7,
    "D+": 1.3, "D": 1.0, "D-": 0.7,
    "F":  0.0,
}


def _parse_transcript_xlsx(data: bytes) -> dict:
    """
    Parse the WashU 'View My Academic Record' xlsx export.
    Returns {"semesters": [...], "transfer_credits": [...]}.
    Each semester: {"term": str, "year": int, "courses": [...]}.
    Each course: {"code", "name", "credits", "grade", "grade_points"}.
    """
    wb = load_workbook(filename=BytesIO(data), data_only=True)
    ws = wb.active

    STUDENT_PREFIX = "Phan,"   # col-A prefix that marks actual course rows

    semesters: list[dict] = []
    transfer_credits: list[dict] = []
    current_sem: dict | None = None
    in_transfer = False

    for row in ws.iter_rows(values_only=True):
        a, b, c, d, e = (row[i] if i < len(row) else None for i in range(5))

        # ── section markers ──────────────────────────────────────────────
        if a == "Transfer Credit from Exams":
            in_transfer = True
            continue

        if a == "Academic Period" and isinstance(b, str) and not in_transfer:
            parts = b.strip().split(" ")           # ["Spring", "2026"]
            if len(parts) == 2 and parts[1].isdigit():
                current_sem = {"term": parts[0], "year": int(parts[1]), "courses": []}
                semesters.append(current_sem)
            continue

        # ── skip non-course rows ─────────────────────────────────────────
        if not isinstance(a, str) or not a.startswith("Phan,"):
            continue

        # ── parse course code + name from col B ──────────────────────────
        if not isinstance(b, str) or " - " not in b:
            continue
        code_part, _, name_part = b.partition(" - ")
        code = code_part.strip()
        name = name_part.strip()

        # ── transfer / AP credits ────────────────────────────────────────
        if in_transfer:
            # col C = credits (int), col D = "Exam Credit"
            credits = float(c) if isinstance(c, (int, float)) else 0.0
            transfer_credits.append({
                "code": code, "name": name,
                "credits": credits, "grade": "CR",
            })
            continue

        # ── regular enrolled course ──────────────────────────────────────
        # col C = grade str, col D = grade_points float|None, col E = credits
        if current_sem is None:
            continue
        grade_raw = str(c).strip() if c is not None else None
        credits = float(e) if isinstance(e, (int, float)) else 0.0

        # Normalise grade
        if grade_raw in ("Withdraw",):
            grade = "W"
            gp = None
        elif grade_raw in ("Pass", "CR#", None):
            grade = grade_raw
            gp = None
        else:
            grade = grade_raw
            gp = float(d) if isinstance(d, (int, float)) else GRADE_POINTS_MAP.get(grade_raw)

        current_sem["courses"].append({
            "code": code, "name": name,
            "credits": credits, "grade": grade, "grade_points": gp,
        })

    wb.close()
    return {"semesters": semesters, "transfer_credits": transfer_credits}


@router.post("/upload-transcript")
def upload_transcript(
    file: UploadFile = File(...),
    current_user_id: str = Depends(get_current_user),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    raw = file.file.read()
    parsed = _parse_transcript_xlsx(raw)

    db = get_db()

    # 1. Wipe all existing user courses
    db.table("user_courses").delete().eq("user_id", current_user_id).execute()

    # 2. Fetch user's semesters (keyed by term+year)
    existing_sems = db.table("semesters").select("*").eq("user_id", current_user_id).execute().data
    sem_map: dict[tuple, int] = {(s["term"], s["year"]): s["id"] for s in existing_sems}

    # 3. Insert courses semester-by-semester
    courses_to_insert: list[dict] = []
    semesters_to_mark_complete: list[int] = []

    for sem in parsed["semesters"]:
        key = (sem["term"], sem["year"])
        if key not in sem_map:
            # Create the semester if it doesn't exist in the user's list
            new_sem = db.table("semesters").insert({
                "term": sem["term"], "year": sem["year"],
                "is_completed": True, "user_id": current_user_id,
            }).execute().data[0]
            sem_map[key] = new_sem["id"]

        sem_id = sem_map[key]
        semesters_to_mark_complete.append(sem_id)

        for c in sem["courses"]:
            courses_to_insert.append({
                "user_id": current_user_id,
                "semester_id": sem_id,
                "course_code": c["code"],
                "course_name": c["name"],
                "credits": c["credits"],
                "grade": c["grade"],
                "grade_points": c["grade_points"],
                "is_planned": False,
                "is_transfer": False,
            })

    # 4. Mark semesters that appear in transcript as completed
    for sem_id in semesters_to_mark_complete:
        db.table("semesters").update({"is_completed": True}).eq("id", sem_id).execute()

    # 5. Insert transfer / AP credits (no semester_id)
    for c in parsed["transfer_credits"]:
        courses_to_insert.append({
            "user_id": current_user_id,
            "semester_id": None,
            "course_code": c["code"],
            "course_name": c["name"],
            "credits": c["credits"],
            "grade": c["grade"],
            "grade_points": None,
            "is_planned": False,
            "is_transfer": True,
        })

    if courses_to_insert:
        db.table("user_courses").insert(courses_to_insert).execute()

    return {
        "imported_courses": len([c for c in courses_to_insert if not c["is_transfer"]]),
        "transfer_credits": len(parsed["transfer_credits"]),
        "semesters_updated": len(semesters_to_mark_complete),
    }
